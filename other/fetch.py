import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Paths
DB_PATH = r"D:\GrewAnalytics\warehouse.duckdb"
OUTPUT_FILE = r"D:\OneDrive - CHIRIPAL RENEWABLE ENERGY PRIVATE LIMITED\Desktop\DB1\LC\sheets\treasurydb.xlsx"

# Tables to export
TABLES = [
    "APP_CONFIG",
    "CAPITAL_STACK",
    "DEBT_MATURITY",
    "FX_RATES",
    "SYSTEM_METADATA",
    "TREASURY_INSIGHTS",
    "YIELD_CURVE"
]

# Connect to DuckDB
con = duckdb.connect(DB_PATH, read_only=True)

# Write each table to a separate worksheet
with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    for table in TABLES:
        print(f"Exporting {table}...")

        df = con.execute(f'SELECT * FROM "{table}"').fetchdf()

        df.to_excel(
            writer,
            sheet_name=table[:31],  # Excel worksheet name limit
            index=False
        )

# Post-format workbook
wb = load_workbook(OUTPUT_FILE)

for ws in wb.worksheets:

    # Freeze header row
    ws.freeze_panes = "A2"

    # Enable autofilter
    ws.auto_filter.ref = ws.dimensions

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto-size columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)

wb.save(OUTPUT_FILE)

con.close()

print(f"\nWorkbook created successfully:\n{OUTPUT_FILE}")